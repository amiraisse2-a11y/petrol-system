"""
PETRO AFRICA — Export Excel Natif
Génère un .xlsx professionnel avec openpyxl :
en-têtes colorés, formatage numérique, multi-onglets, logo simulé.
"""

import io
from datetime import datetime

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    GradientFill,
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel

# ─── Couleurs PETRO AFRICA ───────────────────────────────────────────────────
try:
    from config import EXCEL_COULEUR_HEADER, EXCEL_COULEUR_ACCENT, APP_NAME, APP_VERSION
except ImportError:
    EXCEL_COULEUR_HEADER = "1B4F72"
    EXCEL_COULEUR_ACCENT = "F39C12"
    APP_NAME    = "PETRO AFRICA"
    APP_VERSION = "2.1.0"

# Fills
FILL_HEADER  = PatternFill("solid", fgColor=EXCEL_COULEUR_HEADER)
FILL_ACCENT  = PatternFill("solid", fgColor=EXCEL_COULEUR_ACCENT)
FILL_ALT1    = PatternFill("solid", fgColor="EBF5FB")   # bleu très clair
FILL_ALT2    = PatternFill("solid", fgColor="FDFEFE")   # blanc cassé
FILL_TITLE   = PatternFill("solid", fgColor="0D2137")   # presque noir

# Fonts
FONT_TITLE   = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
FONT_HEADER  = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_SUBHEAD = Font(name="Calibri", size=11, bold=True, color=EXCEL_COULEUR_HEADER)
FONT_BODY    = Font(name="Calibri", size=10)
FONT_BOLD    = Font(name="Calibri", size=10, bold=True)
FONT_SMALL   = Font(name="Calibri", size=9, color="666666")

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_RIGHT  = Alignment(horizontal="right",  vertical="center")
ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center")

BORDER_THIN  = Border(
    bottom=Side(style="thin", color="BDC3C7"),
    right=Side(style="thin", color="BDC3C7"),
)
BORDER_THICK = Border(
    bottom=Side(style="medium", color=EXCEL_COULEUR_HEADER),
)


# ═══════════════════════════════════════════════════════════════════════════════
#  HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def set_col_widths(ws, widths: dict):
    """widths = {'A': 18, 'B': 12, ...}"""
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def write_header_row(ws, row: int, headers: list[str], fill=None):
    fill = fill or FILL_HEADER
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font      = FONT_HEADER
        cell.fill      = fill
        cell.alignment = ALIGN_CENTER
        cell.border    = BORDER_THIN


def write_data_row(ws, row: int, values: list, bold: bool = False,
                   color_neg: bool = False, alt: bool = False):
    fill = FILL_ALT1 if alt else FILL_ALT2
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font   = FONT_BOLD if bold else FONT_BODY
        cell.fill   = fill
        cell.border = BORDER_THIN
        cell.alignment = ALIGN_RIGHT if isinstance(val, (int, float)) else ALIGN_LEFT
        if color_neg and isinstance(val, (int, float)):
            cell.font = Font(name="Calibri", size=10,
                             bold=bold,
                             color="C0392B" if val < 0 else "1E8449")


def write_title_block(ws, title: str, subtitle: str = "", start_row: int = 1):
    """Bloc titre en-tête de feuille."""
    ws.merge_cells(f"A{start_row}:H{start_row}")
    cell = ws.cell(row=start_row, column=1, value=f"  {APP_NAME} — {title}")
    cell.font      = FONT_TITLE
    cell.fill      = FILL_TITLE
    cell.alignment = ALIGN_LEFT
    ws.row_dimensions[start_row].height = 32

    if subtitle:
        ws.merge_cells(f"A{start_row+1}:H{start_row+1}")
        cell2 = ws.cell(row=start_row + 1, column=1,
                        value=f"  {subtitle}  |  Généré le {datetime.now():%d/%m/%Y %H:%M}")
        cell2.font      = FONT_SMALL
        cell2.fill      = FILL_ACCENT
        cell2.alignment = ALIGN_LEFT
        ws.row_dimensions[start_row + 1].height = 18
        return start_row + 2

    return start_row + 1


# ═══════════════════════════════════════════════════════════════════════════════
#  ONGLET : PRODUCTION
# ═══════════════════════════════════════════════════════════════════════════════

def write_production_sheet(ws, df_prod: pd.DataFrame):
    ws.title = "Production"
    ws.sheet_view.showGridLines = False

    next_row = write_title_block(ws, "Données de Production",
                                 "Suivi journalier / mensuel — BOPD, MMscfd, GOR, BSW")
    next_row += 1  # ligne vide

    headers = list(df_prod.columns)
    write_header_row(ws, next_row, headers)
    next_row += 1

    for i, (_, row_data) in enumerate(df_prod.iterrows()):
        write_data_row(ws, next_row, list(row_data), alt=(i % 2 == 0))
        next_row += 1

    # Auto-width approximatif
    for col_idx, col in enumerate(headers, start=1):
        max_len = max(len(str(col)), df_prod[col].astype(str).str.len().max())
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 30)

    # Freeze header
    ws.freeze_panes = f"A{next_row - len(df_prod)}"

    # Graphique BarChart production
    try:
        chart = BarChart()
        chart.type = "col"
        chart.title = "Production Mensuelle (BOPD)"
        chart.style = 10
        chart.height = 12
        chart.width  = 22

        # Cherche colonne huile
        huile_cols = [c for c in headers if "huile" in c.lower() or "bopd" in c.lower() or "oil" in c.lower()]
        if huile_cols:
            col_h = headers.index(huile_cols[0]) + 1
            data_ref = Reference(ws,
                                 min_col=col_h,
                                 min_row=next_row - len(df_prod) - 1,
                                 max_row=next_row - 1)
            chart.add_data(data_ref, titles_from_data=True)
            ws.add_chart(chart, f"A{next_row + 2}")
    except Exception:
        pass  # Silently skip chart if data mismatch


# ═══════════════════════════════════════════════════════════════════════════════
#  ONGLET : FINANCES
# ═══════════════════════════════════════════════════════════════════════════════

def write_finances_sheet(ws, df_fin: pd.DataFrame, kpis: dict = None):
    ws.title = "Finances"
    ws.sheet_view.showGridLines = False

    next_row = write_title_block(ws, "Analyse Financière",
                                 "Cash Flow · NPV · IRR · CAPEX/OPEX")
    # KPI summary
    if kpis:
        next_row += 1
        ws.merge_cells(f"A{next_row}:H{next_row}")
        kpi_str = (
            f"NPV: {kpis.get('npv', 'N/A')} MUSD  |  "
            f"IRR: {kpis.get('irr', 'N/A')}%  |  "
            f"Payback: {kpis.get('payback', 'N/A')} ans  |  "
            f"WACC: {kpis.get('wacc', 'N/A')}%"
        )
        cell = ws.cell(row=next_row, column=1, value=kpi_str)
        cell.font      = FONT_SUBHEAD
        cell.alignment = ALIGN_LEFT
        next_row += 1

    next_row += 1
    headers = list(df_fin.columns)
    write_header_row(ws, next_row, headers)
    next_row += 1

    for i, (_, row_data) in enumerate(df_fin.iterrows()):
        is_total = i == len(df_fin) - 1
        write_data_row(ws, next_row, list(row_data),
                       bold=is_total,
                       color_neg=True,
                       alt=(i % 2 == 0))
        if is_total:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=next_row, column=col_idx).border = BORDER_THICK
        next_row += 1

    for col_idx, col in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    ws.freeze_panes = f"B{next_row - len(df_fin)}"

    # Line chart cash flow cumulé
    try:
        cf_cols = [c for c in headers if "cumulé" in c.lower() or "cumul" in c.lower()]
        if cf_cols:
            chart = LineChart()
            chart.title  = "Cash Flow Cumulé (MUSD)"
            chart.style  = 10
            chart.height = 12
            chart.width  = 22
            col_cf = headers.index(cf_cols[0]) + 1
            data_ref = Reference(ws,
                                 min_col=col_cf,
                                 min_row=next_row - len(df_fin) - 1,
                                 max_row=next_row - 1)
            chart.add_data(data_ref, titles_from_data=True)
            ws.add_chart(chart, f"A{next_row + 2}")
    except Exception:
        pass


# ═══════════════════════════════════════════════════════════════════════════════
#  ONGLET : ESG
# ═══════════════════════════════════════════════════════════════════════════════

def write_esg_sheet(ws, df_esg: pd.DataFrame, score: dict = None):
    ws.title = "ESG & Flaring"
    ws.sheet_view.showGridLines = False

    FILL_ESG = PatternFill("solid", fgColor="0E6655")

    next_row = write_title_block(ws, "Rapport ESG & Torchage",
                                 "CO2 équivalent · Conformité PETROCI · Score ESG")
    if score:
        next_row += 1
        ws.merge_cells(f"A{next_row}:H{next_row}")
        score_str = (
            f"Score ESG: {score.get('score_total', 'N/A')}/100  |  "
            f"Grade: {score.get('grade', 'N/A')}  |  "
            f"Objectif PETROCI 2030: ≤ 2% torchage"
        )
        cell = ws.cell(row=next_row, column=1, value=score_str)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = FILL_ESG
        cell.alignment = ALIGN_LEFT
        next_row += 1

    next_row += 1
    headers = list(df_esg.columns)
    write_header_row(ws, next_row, headers, fill=FILL_ESG)
    next_row += 1

    for i, (_, row_data) in enumerate(df_esg.iterrows()):
        write_data_row(ws, next_row, list(row_data), alt=(i % 2 == 0))

        # Colorer statut
        statut_col = [j for j, h in enumerate(headers) if h == "Statut"]
        if statut_col:
            cell = ws.cell(row=next_row, column=statut_col[0] + 1)
            val  = str(row_data.get("Statut", ""))
            cell.font = Font(
                name="Calibri", size=10,
                color=("1E8449" if "Conforme" in val
                       else "D4AC0D" if "Surveillance" in val
                       else "C0392B")
            )
        next_row += 1

    for col_idx, col in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20


# ═══════════════════════════════════════════════════════════════════════════════
#  ONGLET : RÉSERVOIR
# ═══════════════════════════════════════════════════════════════════════════════

def write_reservoir_sheet(ws, df_res: pd.DataFrame, type_calc: str = "P/z"):
    ws.title = "Réservoir"
    ws.sheet_view.showGridLines = False

    next_row = write_title_block(ws, f"Analyse Réservoir — {type_calc}",
                                 "P/z Plot · IPR · OGIP/OOIP")
    next_row += 1
    headers = list(df_res.columns)
    write_header_row(ws, next_row, headers)
    next_row += 1

    for i, (_, row_data) in enumerate(df_res.iterrows()):
        write_data_row(ws, next_row, list(row_data), alt=(i % 2 == 0))
        next_row += 1

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18


# ═══════════════════════════════════════════════════════════════════════════════
#  ONGLET : ALERTES
# ═══════════════════════════════════════════════════════════════════════════════

def write_alerts_sheet(ws, df_alerts: pd.DataFrame):
    ws.title = "Alertes"
    ws.sheet_view.showGridLines = False

    FILL_DANGER  = PatternFill("solid", fgColor="922B21")
    FILL_WARN    = PatternFill("solid", fgColor="B7770D")

    next_row = write_title_block(ws, "Journal des Alertes",
                                 "Anomalies détectées · Seuils dépassés")
    next_row += 1
    headers = list(df_alerts.columns)
    write_header_row(ws, next_row, headers, fill=FILL_DANGER)
    next_row += 1

    for i, (_, row_data) in enumerate(df_alerts.iterrows()):
        sev = str(row_data.get("Sévérité", row_data.get("severity", ""))).lower()
        alt_fill = (
            PatternFill("solid", fgColor="FDEDEC") if "critique" in sev or "critical" in sev
            else PatternFill("solid", fgColor="FEF9E7") if "warn" in sev or "alerte" in sev
            else FILL_ALT2
        )
        for col_idx, val in enumerate(list(row_data), start=1):
            cell = ws.cell(row=next_row, column=col_idx, value=val)
            cell.font      = FONT_BODY
            cell.fill      = alt_fill
            cell.border    = BORDER_THIN
            cell.alignment = ALIGN_LEFT
        next_row += 1

    for col_idx, col in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 22


# ═══════════════════════════════════════════════════════════════════════════════
#  ONGLET : COUVERTURE (page de garde)
# ═══════════════════════════════════════════════════════════════════════════════

def write_cover_sheet(ws, metadata: dict):
    ws.title = "Couverture"
    ws.sheet_view.showGridLines = False
    ws.sheet_view.showRowColHeaders = False

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 40

    # Titre principal
    ws.merge_cells("B2:G5")
    cell = ws.cell(row=2, column=2, value=APP_NAME)
    cell.font      = Font(name="Calibri", size=32, bold=True, color=EXCEL_COULEUR_HEADER)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.fill      = PatternFill("solid", fgColor="F2F3F4")
    ws.row_dimensions[2].height = 40
    ws.row_dimensions[3].height = 40

    # Sous-titre
    ws.merge_cells("B6:G6")
    sub = ws.cell(row=6, column=2,
                  value="Dashboard Pétrolier — Rapport Export")
    sub.font = Font(name="Calibri", size=14, italic=True,
                    color=EXCEL_COULEUR_ACCENT)

    # Ligne décorative
    for col in range(2, 8):
        cell = ws.cell(row=7, column=col)
        cell.fill = FILL_ACCENT

    # Métadonnées
    infos = [
        ("Rapport généré le",   datetime.now().strftime("%d/%m/%Y à %H:%M")),
        ("Version Dashboard",   APP_VERSION),
        ("Champ / Projet",      metadata.get("champ", "—")),
        ("Opérateur",           metadata.get("operateur", "—")),
        ("Période",             metadata.get("periode", "—")),
        ("Préparé par",         metadata.get("auteur", APP_NAME)),
        ("Confidentialité",     metadata.get("confidentialite", "CONFIDENTIEL")),
    ]
    for i, (label, value) in enumerate(infos, start=9):
        lbl_cell = ws.cell(row=i, column=2, value=label)
        lbl_cell.font      = FONT_BOLD
        lbl_cell.alignment = ALIGN_LEFT

        val_cell = ws.cell(row=i, column=3, value=value)
        val_cell.font      = FONT_BODY
        val_cell.alignment = ALIGN_LEFT

        ws.row_dimensions[i].height = 18

    # Pied de page
    ws.merge_cells("B18:G18")
    footer = ws.cell(row=18, column=2,
                     value=f"© {datetime.now().year} {APP_NAME} — Document confidentiel")
    footer.font      = FONT_SMALL
    footer.fill      = FILL_TITLE
    footer.font      = Font(name="Calibri", size=9, color="AAAAAA")
    footer.alignment = ALIGN_CENTER


# ═══════════════════════════════════════════════════════════════════════════════
#  FONCTION PRINCIPALE D'EXPORT
# ═══════════════════════════════════════════════════════════════════════════════

def generer_rapport_excel(
    dfs: dict,           # {"production": df, "finances": df, "esg": df, ...}
    metadata: dict = None,
    kpis: dict    = None,
    score_esg: dict = None,
) -> bytes:
    """
    Génère un workbook Excel complet.

    Paramètres
    ----------
    dfs : dict de DataFrames, clés possibles :
          "production", "finances", "esg", "reservoir", "alertes"
    metadata : dict pour la page de couverture
    kpis : dict NPV/IRR/Payback pour onglet finances
    score_esg : dict score ESG

    Retourne les bytes du fichier .xlsx.
    """
    metadata = metadata or {}
    wb = Workbook()
    wb.remove(wb.active)   # Supprime la feuille vide par défaut

    # 1. Couverture (toujours en premier)
    ws_cover = wb.create_sheet("Couverture")
    write_cover_sheet(ws_cover, metadata)

    # 2. Onglets données
    SHEET_WRITERS = {
        "production": lambda ws: write_production_sheet(ws, dfs["production"]),
        "finances":   lambda ws: write_finances_sheet(ws, dfs["finances"], kpis),
        "esg":        lambda ws: write_esg_sheet(ws, dfs["esg"], score_esg),
        "reservoir":  lambda ws: write_reservoir_sheet(ws, dfs["reservoir"]),
        "alertes":    lambda ws: write_alerts_sheet(ws, dfs["alertes"]),
    }
    for key, writer in SHEET_WRITERS.items():
        if key in dfs and dfs[key] is not None and len(dfs[key]) > 0:
            ws = wb.create_sheet()
            writer(ws)

    # Sauvegarde en mémoire
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ═══════════════════════════════════════════════════════════════════════════════
#  COMPOSANT STREAMLIT — Bouton téléchargement
# ═══════════════════════════════════════════════════════════════════════════════

def render_export_excel_button(
    dfs: dict,
    metadata: dict = None,
    kpis: dict    = None,
    score_esg: dict = None,
    label: str    = "📥 Télécharger Rapport Excel (.xlsx)",
    filename: str = None,
):
    """
    Affiche un bouton Streamlit de téléchargement Excel.

    Usage :
        from export_excel import render_export_excel_button
        render_export_excel_button(
            dfs={"production": df_prod, "finances": df_fin},
            metadata={"champ": "Baleine", "operateur": "ENI/PETROCI"},
            kpis={"npv": 142.3, "irr": 18.5, "payback": 7.2},
        )
    """
    if not filename:
        ts = datetime.now().strftime("%Y%m%d_%H%M")
        champ = (metadata or {}).get("champ", "RAPPORT")
        filename = f"PETRO_AFRICA_{champ}_{ts}.xlsx"

    try:
        excel_bytes = generer_rapport_excel(dfs, metadata, kpis, score_esg)
        st.download_button(
            label=label,
            data=excel_bytes,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as e:
        st.error(f"Erreur génération Excel : {e}")


# ═══════════════════════════════════════════════════════════════════════════════
#  DEMO STANDALONE
# ═══════════════════════════════════════════════════════════════════════════════

def _demo():
    """Démonstration avec données synthétiques."""
    import numpy as np

    st.title("📊 Demo Export Excel PETRO AFRICA")

    n = 12
    df_prod = pd.DataFrame({
        "Mois":           pd.date_range("2024-01", periods=n, freq="M").strftime("%b %Y"),
        "Prod. Huile (BOPD)": np.random.randint(4000, 7000, n),
        "Prod. Gaz (MMscfd)": np.round(np.random.uniform(15, 25, n), 2),
        "Water Cut (%)":  np.round(np.random.uniform(15, 40, n), 1),
        "GOR (scf/bbl)":  np.random.randint(400, 800, n),
    })
    df_fin = pd.DataFrame({
        "Année":               range(1, 6),
        "Revenus (MUSD)":      [42.0, 38.5, 35.2, 32.1, 29.4],
        "CAPEX (MUSD)":        [120.0, 50.0, 10.0, 5.0, 0.0],
        "OPEX (MUSD)":         [18.0, 17.0, 16.0, 15.0, 14.0],
        "Cash Flow Net (MUSD)":[-96.0, -28.5, 9.2, 12.1, 15.4],
        "Cash Flow Cumulé (MUSD)": [-96.0, -124.5, -115.3, -103.2, -87.8],
    })
    df_esg = pd.DataFrame({
        "Année":               [2025, 2026, 2027, 2028, 2029, 2030],
        "Flaring (%)":         [5.0, 4.2, 3.5, 3.0, 2.4, 2.0],
        "CO2e Annuel (ktCO2)": [12.5, 10.3, 8.7, 7.4, 6.0, 5.1],
        "Valeur Perdue (kUSD)": [225, 188, 156, 132, 106, 90],
        "Statut":              ["⚠️ Surveillance","⚠️ Surveillance","⚠️ Surveillance",
                                "⚠️ Surveillance","✅ Conforme","✅ Conforme"],
    })

    render_export_excel_button(
        dfs={"production": df_prod, "finances": df_fin, "esg": df_esg},
        metadata={"champ": "Baleine", "operateur": "ENI/PETROCI",
                  "periode": "2024–2029", "auteur": "Ingénieur Production"},
        kpis={"npv": "−87.8", "irr": "N/A", "payback": "Hors horizon"},
    )
    st.info("Cliquez le bouton ci-dessus pour télécharger le rapport Excel formaté.")


if __name__ == "__main__":
    _demo()
